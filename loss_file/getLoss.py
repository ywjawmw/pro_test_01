# -*- coding: utf-8 -*- 
# @Time : 2021/6/24 10:12 
# @Author : Ywj 
# @File : getLoss.py 
# @function :  统计loss
from openpyxl import load_workbook

def readLoss(file_name):
    print("分割herbPair-40 file !!!")
    print('***************开始载入文件！！！************')
    lines = open(file_name, 'r', encoding="utf-8").readlines()
    loss = []
    link_loss = []
    for line in lines:
        if "train==[" in line:
            line = line[line.find("train==[") + len("train==["):]
            line1 = line[:line.find("=")]
            loss.append(line1)
            link = line[line.find(" + ") + len(" + "):]
            link = link[:link.find(" + ")]
            link_loss.append(link)
    print(len(link_loss))
    return loss, link_loss


def saveLoss(filename, loss):
    print("写入Excel文件")
    file = load_workbook(filename)  # 生成一个已存在的wookbook对象
    save_file = file.active
    for i in range(2, 2002):
        save_file.cell(i, 3, loss[i-2])
    file.save(filename)


if __name__ == '__main__':
    loss_filepath = r"loss.txt"
    save_filepath = r"loss.xlsx"
    loss, link = readLoss(loss_filepath)
    saveLoss(save_filepath, link)
